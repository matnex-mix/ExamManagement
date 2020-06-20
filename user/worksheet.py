import openpyxl
from openpyxl.styles import Font
import os
def CompileResult(IndividualSheet, CompiledSheet,CourseCode,CourseUnit,Session):
    # sheet to be added to the compiled sheet
    ResultWb = openpyxl.load_workbook(IndividualSheet)
    wb = ''
    sheet =''
    ResultSheet = ''
    # try to check if the file exist before
    try:
        wb = openpyxl.load_workbook(CompiledSheet)
        sheet = wb.active
        ResultSheet = ResultWb.active
    # if the file is not created before
    except:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.merge_cells('A1:M1')
        sheet['A1'] = 'HND 1 COMPUTER SCIENCE, '+ Session+ ' '+CourseCode
        font = Font(name='Calibri' ,size= 20, bold=True )
        sheet['A1'].font = font
        sheet.column_dimensions['B'].width = 20
        sheet['A2'] = 'Units'
        sheet['A3'] = 'S/N'
        sheet['B3'] = 'MATRIC NUMBER'
        sn = 1
        ResultSheet = ResultWb.active
        for i in range(4, ResultSheet.max_row + 3):
            sheet.cell(row=i, column=1).value = sn
            sheet.cell(row=i, column=2).value = ResultSheet.cell(row=i-2, column=1).value
            sn+=1
        sheet['N1'] = 'TOTAL POINTS'
        sheet.column_dimensions['N'].width = 20
        sheet['O1'] = 'ALL UNITS'
        sheet.column_dimensions['O'].width = 20
        sheet['P1'] = 'GP'
        sheet.column_dimensions['P'].width = 20
        wb.save(CompiledSheet)
    Column =  ''
    # loop to get the next empty column
    for i in range(3, sheet.max_column):
        if sheet.cell(row=3, column=i).value == None:
            Column = i
            break
    # this line add the course unit
    sheet.cell(row=2,column=Column).value= CourseUnit
    # this line add the course Course code
    sheet.cell(row=3,column=Column).value= CourseCode
    """
    This loop populates the sheet with data from the forwarded sheet
    """
    for i in range(4, ResultSheet.max_row + 3):
        grade = ResultSheet.cell(row=i -2, column=3).value
        # this update grade to the sheet
        sheet.cell(row=i, column=Column).value = grade
        unit = sheet.cell(row = 2, column = Column).value
        # get all total units calculated before
        InitialTotalUnits =sheet.cell(row=i, column=15).value
        # get all total points calculated before
        InitialTotalPoints = sheet.cell(row=i, column=14).value
        if InitialTotalUnits != None:
            sheet.cell(row=i, column=15).value += unit
        else:
            sheet.cell(row=i, column=15).value = unit
        if InitialTotalPoints == None:
            sheet.cell(row=i, column=14).value = 0
        if grade == 'A':
            sheet.cell(row=i, column=14).value += unit * 4
        elif grade == 'AB':
            sheet.cell(row=i, column=14).value += unit * 3.5
        elif grade == 'B':
            sheet.cell(row=i, column=14).value += unit * 3.25
        elif grade == 'BC':
            sheet.cell(row=i, column=14).value += unit * 3
        elif grade == 'C':
            sheet.cell(row=i, column=14).value += unit * 2.75
        elif grade == 'CD':
            sheet.cell(row=i, column=14).value += unit * 2.50
        elif grade == 'D':
            sheet.cell(row=i, column=14).value += unit * 2.25
        elif grade == 'E':
            sheet.cell(row=i, column=14).value += unit * 2.00
        else:
            sheet.cell(row=i, column=14).value += unit * 0
        sheet.cell(row=i, column=16).value = round(sheet.cell(row=i, column=14).value / sheet.cell(row=i, column=15).value,2)
    wb.save(CompiledSheet)
